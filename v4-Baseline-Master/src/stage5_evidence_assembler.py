import os
import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.comments import Comment

# Configuration
PROJECT_ROOT = Path("/Users/joshuabhawanlall/Git Folder/Clinical_AI_Extraction_Pipeline")
MERGED_DIR = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/longitudinal_merged"
HARVEST_DIR = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/raw_harvest"
PROTOTYPE_WORKBOOK = PROJECT_ROOT / "data/hackathon-database-prototype.xlsx"
OUTPUT_WORKBOOK = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/generated-database-v7-diamond.xlsx"

def find_evidence(value):
    # Search all harvest files for this value's evidence
    for f in os.listdir(HARVEST_DIR):
        with open(HARVEST_DIR / f) as j:
            h = json.load(j)
            for c in h["candidates"]:
                if str(c.get("value")) == str(value):
                    return c.get("evidence", "No snippet found.")
    return "Consolidated record."

def main():
    if not os.path.exists(OUTPUT_WORKBOOK.parent): os.makedirs(OUTPUT_WORKBOOK.parent)
    merged_files = sorted([f for f in os.listdir(MERGED_DIR) if f.endswith(".json")])
    
    final_rows = []
    for f in merged_files:
        with open(MERGED_DIR / f) as j: final_rows.append(json.load(j))
            
    df = pd.DataFrame(final_rows)
    template_cols = list(pd.read_excel(PROTOTYPE_WORKBOOK).columns)
    df = df.reindex(columns=template_cols)
    
    wb = load_workbook(PROTOTYPE_WORKBOOK)
    ws = wb.active
    if ws.max_row >= 2: ws.delete_rows(2, ws.max_row)

    print("Assembling Diamond V1 with Evidence Comments...")
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for c_idx, value in enumerate(row.tolist(), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            val_str = str(value).strip() if value and str(value).lower() != "nan" else ""
            if val_str.endswith(".0"): val_str = val_str[:-2]
            
            if val_str:
                cell.value = val_str
                # Evidence
                evidence = find_evidence(val_str)
                cell.comment = Comment(f"AI Evidence:\n{evidence}", "Gemini CLI")
            
            cell.number_format = '@'
            cell.data_type = 's'
            
    wb.save(OUTPUT_WORKBOOK)
    print(f"Diamond V1 Assembly Complete.")

if __name__ == "__main__":
    main()
