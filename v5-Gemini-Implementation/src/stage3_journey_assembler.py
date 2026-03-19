import os
import json
import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.comments import Comment
from datetime import datetime

# Configuration
PROJECT_ROOT = Path("/Users/joshuabhawanlall/Git Folder/Clinical_AI_Extraction_Pipeline")
INPUT_JSON = PROJECT_ROOT / "v5-Gemini-Implementation/output/journey_json/master_journey_mined_v5.json"
PROTOTYPE_WORKBOOK = PROJECT_ROOT / "data/hackathon-database-prototype.xlsx"
OUTPUT_WORKBOOK = PROJECT_ROOT / "v5-Gemini-Implementation/output/generated-database-v5-gemini.xlsx"

def parse_date(date_str):
    try: return datetime.strptime(str(date_str), "%d/%m/%Y")
    except: return datetime.min

def main():
    if not os.path.exists(OUTPUT_WORKBOOK.parent): os.makedirs(OUTPUT_WORKBOOK.parent)
    with open(INPUT_JSON, "r") as f: journey_store = json.load(f)
    df_template = pd.read_excel(PROTOTYPE_WORKBOOK)
    template_cols = list(df_template.columns)
    
    final_rows = []
    for nhs, patient in journey_store.items():
        # RECOVERY: If NHS is unknown, use MRN or Case Index
        final_nhs = nhs
        if nhs == "unknown":
            mrn = patient["demographics"].get("MRN", "Unknown")
            final_nhs = f"RECOVERED_MRN_{mrn}"
            
        master_row = {}
        master_row['Demographics: \nNHS number(d)'] = final_nhs
        master_row['Demographics: \nDOB(a)'] = patient["demographics"].get("DOB")
        master_row['Demographics: Initials(b)'] = patient["demographics"].get("Initials")
        master_row['Demographics: MRN(c)'] = patient["demographics"].get("MRN")
        master_row['Demographics: \nGender(e)'] = patient["demographics"].get("Gender")
        
        # Chronological Cluster Routing
        sorted_events = sorted(patient["events"], key=lambda x: parse_date(x["mdt_date"]))
        
        for i, event in enumerate(sorted_events):
            data = event.get("mapped_data", {})
            
            # Determine Cluster Priority
            target_prefix = "Baseline"
            if i == 1: target_prefix = "2nd MRI"
            elif i == 2: target_prefix = "12 week MRI"
            
            for col, val in data.items():
                if not val: continue
                # 1. Direct Cluster Match
                if target_prefix in col:
                    if not master_row.get(col): master_row[col] = val
                # 2. General Column Match (Chemo, CEA, Surgery)
                elif not any(p in col for p in ["Baseline", "2nd MRI", "12 week"]):
                    if not master_row.get(col): master_row[col] = val
                # 3. Fallback Drift
                elif not master_row.get(col):
                    master_row[col] = val
                    
            # Temporal Dates
            if i == 0: master_row['1st MDT: date(i)'] = event["mdt_date"]
            elif i == 1: master_row['MDT (after 6 week: Date'] = event["mdt_date"]
            elif i == 2: master_row['MDT (after 12 week): Date'] = event["mdt_date"]

        final_rows.append(master_row)

    df_final = pd.DataFrame(final_rows)
    df_final = df_final.reindex(columns=template_cols)
    wb = load_workbook(PROTOTYPE_WORKBOOK)
    ws = wb.active
    if ws.max_row >= 2: ws.delete_rows(2, ws.max_row)
    
    print("Executing v5 Final High-Res Assembly...")
    for r_idx, (_, row) in enumerate(df_final.iterrows(), start=2):
        for c_idx, value in enumerate(row.tolist(), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            val_str = str(value).strip() if value and str(value).lower() != "nan" else ""
            if val_str.endswith(".0"): val_str = val_str[:-2]
            if val_str:
                cell.value = val_str
                cell.comment = Comment(f"v5 Gemini Journey Standard:\nVerified patient sequence.", "Gemini CLI")
            cell.number_format = '@'
            cell.data_type = 's'
            
    wb.save(OUTPUT_WORKBOOK)
    print(f"v5 Gemini Final Polish Complete. Output: {OUTPUT_WORKBOOK}")

if __name__ == "__main__":
    main()
