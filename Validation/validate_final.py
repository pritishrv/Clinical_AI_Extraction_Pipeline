"""
validate_final.py — Final Validation Framework
================================================
Validates an extracted Excel output against the Word document ground truth
following the Final Validation methodology:

  - Accuracy = Correct / (Correct + Wrong + Hallucination)
  - Missing data excluded from accuracy, reported as coverage
  - Field-type-appropriate matching thresholds:
      structured  → fuzzy ≥ 85
      clinical    → fuzzy ≥ 85
      freetext    → permissive (substring, stripped prefix, key-term overlap)
  - Two-stage validation: Checkpoint A (Doc→JSON), Checkpoint B (Doc→Excel)
  - Error attribution: Stage 1 vs Stage 2
  - Hallucination detection per group

Usage:
  python validate_final.py \\
    --docx  data/hackathon-mdt-outcome-proformas.docx \\
    --excel path/to/output.xlsx \\
    --jsons v5_VLM_LLM_DirectLogic/output/json_raw_claude/

Dependencies: python-docx, openpyxl, difflib (stdlib)
"""

import argparse
import re
import json
import pathlib
from difflib import SequenceMatcher

import openpyxl
import docx

# ── Matching helpers ──────────────────────────────────────────────────────────

def clean(v):
    if v is None:
        return ""
    return re.sub(r'\s+', ' ', re.sub(r'\([a-zA-Z]\)', '', str(v))).strip()

def norm_date(v):
    s = clean(v)
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
    return f"{m.group(3)}/{m.group(2)}/{m.group(1)}" if m else s

def fuzzy(a, b):
    return SequenceMatcher(
        None,
        " ".join(sorted(a.lower().split())),
        " ".join(sorted(b.lower().split()))
    ).ratio() * 100

def is_match_structured(a, b, is_date=False):
    """Strict match for structured/clinical fields. Threshold ≥ 85."""
    g = norm_date(a) if is_date else clean(a).lower()
    e = norm_date(b) if is_date else clean(b).lower()
    if not g and not e:
        return True
    if not g or not e:
        return False
    return g == e or fuzzy(g, e) >= 85

def is_match_freetext(gt_val, ex_val):
    """
    Permissive match for free-text narrative fields (MDT approach, endoscopy findings).

    Correct if ANY of:
      1. Fuzzy ≥ 70 on full strings
      2. Fuzzy ≥ 85 after stripping filler prefixes (For/Consider/Plan/To/Recommend)
      3. Ground truth is a substring of the Excel value
      4. ≥ 80% of GT key clinical terms appear in the Excel value

    Rationale: MDT outcome fields are narrative prose. 'For neoadjuvant CRT'
    and 'neoadjuvant CRT' record the same clinical decision. Penalising the
    pipeline for including clinical context words is not clinically meaningful.
    """
    g = clean(gt_val).lower()
    e = clean(ex_val).lower()
    if not g and not e:
        return True
    if not g or not e:
        return False

    # 1. Fuzzy ≥ 70 on full strings
    if fuzzy(g, e) >= 70:
        return True

    # 2. Strip filler prefixes and retry with ≥ 85
    FILLER = r'^(?:for|consider|plan|to|recommend(?:ed)?|we will|patient for)\s+'
    g2 = re.sub(FILLER, '', g, flags=re.I).strip()
    e2 = re.sub(FILLER, '', e, flags=re.I).strip()
    if fuzzy(g2, e2) >= 85:
        return True

    # 3. GT contained as substring in Excel value
    if g2 in e or g in e:
        return True

    # 4. ≥ 80% of GT key clinical terms present in Excel value
    STOPWORDS = {
        'with', 'then', 'and', 'for', 'the', 'that', 'this',
        'from', 'into', 'after', 'also', 'were', 'been', 'have',
        'their', 'they', 'which'
    }
    gt_terms = [w for w in re.findall(r'\b\w{4,}\b', g2) if w not in STOPWORDS]
    if gt_terms and sum(1 for w in gt_terms if w in e) / len(gt_terms) >= 0.8:
        return True

    return False

def traceable(val, source_text):
    """Check whether a value can be traced back to the source document."""
    v = clean(val).lower()
    if not v or len(v) < 3:
        return True
    s = clean(source_text).lower()
    if v in s:
        return True
    words = [w for w in v.split() if len(w) > 3]
    if not words:
        return True
    return sum(1 for w in words if w in s) / len(words) >= 0.6

def get_nested(data, path):
    val = data
    for part in path.split('.'):
        val = val.get(part) if isinstance(val, dict) else None
    return val

def do_match(gt_val, ex_val, match_type):
    if match_type == 'date':
        return is_match_structured(gt_val, ex_val, is_date=True)
    if match_type == 'freetext':
        return is_match_freetext(gt_val, ex_val)
    return is_match_structured(gt_val, ex_val, is_date=False)


# ── Ground truth parser ───────────────────────────────────────────────────────

def parse_ground_truth(doc):
    """
    Extract ground truth values from the Word document.
    Returns a list of dicts, one per patient, in document order.
    """
    # Get MDT dates from document body paragraphs (before each table)
    current_mdt = ""
    mdt_dates_ordered = []
    for el in doc.element.body:
        tag = el.tag.split('}')[-1]
        if tag == 'p':
            txt = ''.join(el.itertext())
            m = re.search(
                r'Multidisciplinary Meeting\s+(\d{1,2}[/\-.]\d{1,2}[/\-.]\d{4})', txt
            )
            if m:
                current_mdt = m.group(1).replace('-', '/').replace('.', '/')
        elif tag == 'tbl':
            mdt_dates_ordered.append(current_mdt)

    gt_patients = []
    tidx = 0

    for table in doc.tables:
        if len(table.rows) < 2:
            continue
        if 'patient details' not in ' '.join(c.text for c in table.rows[0].cells).lower():
            continue

        details   = table.rows[1].cells[0].text if len(table.rows) > 1 else ''
        diag_cell = table.rows[3].cells[0].text if len(table.rows) > 3 else ''
        clinical  = table.rows[5].cells[0].text if len(table.rows) > 5 else ''
        mdt_out   = table.rows[7].cells[0].text if len(table.rows) > 7 else ''
        full      = details + '\n' + diag_cell + '\n' + clinical + '\n' + mdt_out

        g = {'_full': full}

        # Demographics
        m = re.search(r'(?:Hospital Number|MRN)[:\s]*([^\n]+)', details)
        if m:
            g['MRN'] = re.sub(r'\([a-zA-Z]\)', '', m.group(1)).strip()
        m = re.search(r'NHS Number[:\s]*([^\n]+)', details)
        if m:
            g['NHS'] = re.sub(r'\([a-zA-Z]\)', '', m.group(1)).strip()
        m = re.search(r'DOB[:\s]*(\d{1,2}[/\-.]\d{1,2}[/\-.]\d{4})', details)
        if m:
            g['DOB'] = m.group(1)
        m = re.search(r'\b(Male|Female)\b', details, re.I)
        if m:
            g['Gender'] = m.group(1)
        if tidx < len(mdt_dates_ordered):
            g['MDT_date'] = mdt_dates_ordered[tidx]

        # Histology
        m = re.search(r'Diagnosis[:\s]*(.+?)(?:\n|ICD|$)', diag_cell, re.I)
        if m:
            v = m.group(1).strip()
            if v:
                g['Diagnosis'] = v
        if re.search(r'deficient MMR|dMMR', full, re.I):
            g['MMR'] = 'Deficient'
        elif re.search(r'proficient MMR|pMMR', full, re.I):
            g['MMR'] = 'Proficient'

        # Endoscopy
        for line in (clinical + '\n' + mdt_out).split('\n'):
            if not re.search(r'\b(Colonoscopy|Flexi\s*sig)\b', line, re.I):
                continue
            if re.search(
                r'(?:for|schedule|repeat|completion|refer\s+for|'
                r'review\s+colonoscopy|for\s+MRI.*and)\s*(?:colonoscopy|flexi)',
                line, re.I
            ):
                continue
            fm = re.search(
                r'(?:Colonoscopy|Flexi\s*sig)[^\n]*?[:\-\u2013]\s*(.{5,})', line, re.I
            )
            remainder = re.sub(
                r'^(?:Colonoscopy|Flexi\s*sig(?:moidoscopy)?)\s*', '', line, flags=re.I
            ).strip()
            findings = (
                fm.group(1).strip().rstrip('.,;') if fm
                else (remainder.rstrip('.,;') if len(remainder) >= 5 else None)
            )
            if findings:
                g['Endo_findings'] = findings
                dm = re.search(
                    r'(?:Colonoscopy|Flexi\s*sig)\s+(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})',
                    line, re.I
                )
                if dm:
                    g['Endo_date'] = dm.group(1)
                if re.search(r'flexi\s*sig', line, re.I):
                    g['Endo_type'] = (
                        'incomplete colonoscopy'
                        if re.search(r'unable to reach', line, re.I)
                        else 'flexi sig'
                    )
                else:
                    g['Endo_type'] = (
                        'colonoscopy complete'
                        if re.search(r'ileoc[ae]cal|to TI\b', line, re.I)
                        else 'colonoscopy'
                    )
                break

        # Baseline CT
        for pat in [
            re.compile(
                r'\bCT(?:\s+(?:TAP|abdomen|pelvis|colonography|chest)|\s+\d)'
                r'[^\n]*?(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})[^\n]*', re.I
            ),
            re.compile(r'\bCT\s+(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})[^\n]*', re.I),
        ]:
            m = pat.search(full)
            if m:
                g['CT_date'] = m.group(1).replace('-', '/')
                line = m.group(0)
                t    = re.search(r'\bT(\d[a-d]?)\b', line, re.I)
                n    = re.search(r'\bN(\d[a-c]?)\b', line)
                ms   = re.search(r'\bM(\d)\b', line)
                emvi = re.search(r'EMVI\s*(positive|negative)', line, re.I)
                if t:    g['CT_T']    = 'T' + t.group(1)
                if n:    g['CT_N']    = 'N' + n.group(1)
                if ms:   g['CT_M']    = 'M' + ms.group(1)
                if emvi: g['CT_EMVI'] = emvi.group(1).lower()
                if re.search(r'incidental|adrenal|lung\s*nodule', line, re.I):
                    g['CT_incidental'] = 'Y'
                break

        # Baseline MRI
        for pat in [
            re.compile(
                r'\bMRI\s+(?:pelvis|rectum)[^\n]*?(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})[^\n]*',
                re.I
            ),
            re.compile(r'\bMRI\s+(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})[^\n]*', re.I),
            re.compile(r'\bMRI\s+stage[:\s]+([^\n]+)', re.I),
        ]:
            m = pat.search(full)
            if m:
                line = m.group(0)
                d    = re.search(r'(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})', line)
                if d: g['MRI_date'] = d.group(1).replace('-', '/')
                t    = re.search(r'\bT(\d[a-d]?)\b', line, re.I)
                n    = re.search(r'\bN(\d[a-c]?)\b', line, re.I)
                emvi = re.search(r'EMVI\s*(positive|negative)', line, re.I)
                crm  = re.search(r'CRM\s*(clear|threatened|involved|unsafe)', line, re.I)
                psw  = re.search(r'PSW\s*(clear|negative|positive)', line, re.I)
                if t:    g['MRI_T']    = 'T' + t.group(1)
                if n:    g['MRI_N']    = 'N' + n.group(1)
                if emvi: g['MRI_EMVI'] = emvi.group(1).lower()
                if crm:
                    v = crm.group(1).lower()
                    g['MRI_CRM'] = 'involved' if 'unsafe' in v else v
                if psw:  g['MRI_PSW']  = psw.group(1).lower()
                break

        # MDT treatment approach
        mdt_m = re.search(
            r'(?:TNT|neoadjuvant|CRT|straight\s+to\s+surgery|downstaging|Papillon|EBRT|'
            r'short.?course|FOXTROT|watch\s+and\s+wait|local\s+excision|hemicolectomy|'
            r'resection|colectomy|ESD|refer\s+for\s+colorectal|surgical\s+review|'
            r'For\s+(?:right|left)|right\s+hemicolectomy|left\s+hemicolectomy)[^\n]*',
            mdt_out, re.I
        )
        if mdt_m:
            g['MDT_approach'] = mdt_m.group(0).strip()

        gt_patients.append(g)
        tidx += 1

    return gt_patients


# ── Field and group definitions ───────────────────────────────────────────────

# (gt_key, excel_column_header, json_path, match_type)
# match_type: 'structured' | 'date' | 'clinical' | 'freetext'
FIELDS = [
    ('MRN',         'Demographics: MRN(c)',
     'patient_details.hospital_number', 'structured'),
    ('NHS',         'Demographics: \nNHS number(d)',
     'patient_details.nhs_number', 'structured'),
    ('DOB',         'Demographics: \nDOB(a)',
     'patient_details.dob', 'date'),
    ('Gender',      'Demographics: \nGender(e)',
     'patient_details.gender', 'structured'),
    ('MDT_date',    '1st MDT: date(i)',
     'mdt_meeting_date', 'date'),
    ('Diagnosis',   'Histology: Biopsy result(g)',
     'staging_and_diagnosis.diagnosis', 'clinical'),
    ('MMR',         'Histology: \nMMR status(g/h)',
     'staging_and_diagnosis.mmr_status', 'clinical'),
    ('Endo_date',   'Endoscopy: date(f)',
     'endoscopy.date', 'date'),
    ('Endo_type',
     'Endosopy type: flexi sig, incomplete colonoscopy, colonoscopy complete'
     ' - if gets to ileocecal valve(f) ',
     'endoscopy.type', 'clinical'),
    ('Endo_findings', 'Endoscopy: Findings(f)',
     'endoscopy.findings', 'freetext'),
    ('CT_date',     'Baseline CT: Date(h)',
     'baseline_ct.date', 'date'),
    ('CT_T',        'Baseline CT: T(h)',
     'baseline_ct.T', 'structured'),
    ('CT_N',        'Baseline CT: N(h)',
     'baseline_ct.N', 'structured'),
    ('CT_M',        'Baseline CT: M(h)',
     'baseline_ct.M', 'structured'),
    ('CT_EMVI',     'Baseline CT: EMVI(h)',
     'baseline_ct.EMVI', 'structured'),
    ('MRI_date',    'Baseline MRI: date(h)',
     'baseline_mri.date', 'date'),
    ('MRI_T',       'Baseline MRI: mrT(h)',
     'baseline_mri.mrT', 'structured'),
    ('MRI_N',       'Baseline MRI: mrN(h)',
     'baseline_mri.mrN', 'structured'),
    ('MRI_EMVI',    'Baseline MRI: mrEMVI(h)',
     'baseline_mri.mrEMVI', 'structured'),
    ('MRI_CRM',     'Baseline MRI: mrCRM(h)',
     'baseline_mri.mrCRM', 'structured'),
    ('MRI_PSW',     'Baseline MRI: mrPSW(h)',
     'baseline_mri.mrPSW', 'structured'),
    ('MDT_approach',
     '1st MDT: Treatment approach \n(TNT, downstaging chemotherapy, downstaging nCRT,'
     ' downstaging shortcourse RT, Papillon +/- EBRT, straight to surgery(h)',
     'mdt_outcome', 'freetext'),
]

GROUPS = {
    'Demographics': ['MRN', 'NHS', 'DOB', 'Gender', 'MDT_date'],
    'Histology':    ['Diagnosis', 'MMR'],
    'Endoscopy':    ['Endo_date', 'Endo_type', 'Endo_findings'],
    'Baseline CT':  ['CT_date', 'CT_T', 'CT_N', 'CT_M', 'CT_EMVI'],
    'Baseline MRI': ['MRI_date', 'MRI_T', 'MRI_N', 'MRI_EMVI', 'MRI_CRM', 'MRI_PSW'],
    '1st MDT':      ['MDT_approach'],
}


# ── Scoring ───────────────────────────────────────────────────────────────────

def score_field(gt_patients, excel_rows, jsons, gt_key, excel_col, json_path, match_type):
    """Score a single field across all 50 patients at both checkpoints."""
    A    = {'correct': 0, 'wrong': 0, 'halluc': 0, 'missing': 0, 'na': 0}
    B    = {'correct': 0, 'wrong': 0, 'halluc': 0, 'missing': 0, 'na': 0}
    attr = {'e2e_correct': 0, 'stage2_err': 0, 'stage1_err': 0, 'stage2_fix': 0}

    for i in range(min(50, len(gt_patients), len(excel_rows))):
        gt     = gt_patients[i]
        ex_val = excel_rows[i].get(excel_col)
        js_val = get_nested(jsons[i], json_path) if jsons[i] else None
        gt_val = gt.get(gt_key)
        full   = gt.get('_full', '')

        gt_c = clean(gt_val)
        ex_c = clean(ex_val)
        js_c = clean(js_val)

        # ── Checkpoint A: JSON vs Doc ──────────────────────────────────────
        if not gt_c:
            A['halluc' if js_c else 'na'] += 1
        elif not js_c:
            A['missing'] += 1
        elif do_match(gt_val, js_val, match_type):
            A['correct'] += 1
        else:
            A['wrong'] += 1

        # ── Checkpoint B: Excel vs Doc ─────────────────────────────────────
        if not gt_c:
            if ex_c and not traceable(ex_val, full):
                B['halluc'] += 1
            else:
                B['na'] += 1
        elif not ex_c:
            B['missing'] += 1          # coverage gap — NOT an accuracy failure
        elif do_match(gt_val, ex_val, match_type):
            B['correct'] += 1
        else:
            B['halluc' if not traceable(ex_val, full) else 'wrong'] += 1

        # ── Error attribution ──────────────────────────────────────────────
        json_ok  = bool(gt_c and js_c and do_match(gt_val, js_val, match_type))
        excel_ok = bool(gt_c and ex_c and do_match(gt_val, ex_val, match_type))
        if gt_c:
            if json_ok and excel_ok:
                attr['e2e_correct'] += 1
            elif json_ok and not excel_ok:
                attr['stage2_err'] += 1
            elif not json_ok and not excel_ok:
                attr['stage1_err'] += 1
            elif not json_ok and excel_ok:
                attr['stage2_fix'] += 1

    def acc(d):
        dn = d['correct'] + d['wrong'] + d['halluc']
        return d['correct'] / dn * 100 if dn > 0 else None

    def cov(d):
        pop = d['correct'] + d['wrong'] + d['halluc']
        src = pop + d['missing']
        return pop / src * 100 if src > 0 else None

    return {
        'A': A, 'B': B, 'attr': attr,
        'acc_A': acc(A), 'acc_B': acc(B),
        'cov_A': cov(A), 'cov_B': cov(B),
        'match_type': match_type,
    }


# ── Report printer ────────────────────────────────────────────────────────────

def print_report(results, group_summaries, attr_total):
    SEP  = "=" * 88
    DASH = "─" * 88

    print(f"\n{SEP}")
    print("  FINAL VALIDATION REPORT")
    print("  Accuracy = Correct / (Correct + Wrong + Hallucination)  [missing excluded]")
    print("  Free-text fields use permissive matching (MDT approach, endoscopy findings)")
    print(f"{SEP}")

    for group, fields in GROUPS.items():
        s = group_summaries[group]
        gB = s['B']
        print(f"\n{DASH}")
        print(f"  GROUP: {group.upper()}")
        print(f"{DASH}")
        print(f"  {'Field':<20} {'Type':<12} {'[A]Acc':>7} {'[A]Cov':>7} "
              f"{'[B]Acc':>7} {'[B]Cov':>7}  \u2705  \u274c  \U0001f6a8  \u2b1c")
        print(f"  {'-'*20} {'-'*12} {'-'*7} {'-'*7} {'-'*7} {'-'*7}  --  --  --  --")

        for fk in fields:
            r  = results[fk]
            B  = r['B']; A = r['A']
            aA = f"{r['acc_A']:.0f}%" if r['acc_A'] is not None else "N/A"
            aB = f"{r['acc_B']:.0f}%" if r['acc_B'] is not None else "N/A"
            cA = f"{r['cov_A']:.0f}%" if r['cov_A'] is not None else "N/A"
            cB = f"{r['cov_B']:.0f}%" if r['cov_B'] is not None else "N/A"
            print(f"  {fk:<20} {r['match_type']:<12} {aA:>7} {cA:>7} "
                  f"{aB:>7} {cB:>7}  {B['correct']:>2}  {B['wrong']:>2}  "
                  f"{B['halluc']:>2}  {B['missing']:>2}")

        print(f"\n  {group} \u2192 Stage1: acc={s['acc_A']:.1f}% cov={s['cov_A']:.1f}%  |  "
              f"End-to-end: acc={s['acc_B']:.1f}% cov={s['cov_B']:.1f}%")
        print(f"             \u2705{gB['correct']}  \u274c{gB['wrong']}  "
              f"\U0001f6a8{gB['halluc']}  \u2b1c{gB['missing']}")

    # Category and overall
    def cat_acc(glist):
        c = w = h = 0
        for g in glist:
            b = group_summaries[g]['B']
            c += b['correct']; w += b['wrong']; h += b['halluc']
        d = c + w + h
        return c / d * 100 if d > 0 else 0

    def cat_cov(glist):
        pop = src = 0
        for g in glist:
            b = group_summaries[g]['B']
            pop += b['correct'] + b['wrong'] + b['halluc']
            src += group_summaries[g]['src_B']
        return pop / src * 100 if src > 0 else 0

    demo_acc = cat_acc(['Demographics'])
    demo_cov = cat_cov(['Demographics'])
    clin_g   = ['Histology', 'Endoscopy', 'Baseline CT', 'Baseline MRI', '1st MDT']
    clin_acc = cat_acc(clin_g)
    clin_cov = cat_cov(clin_g)
    overall  = demo_acc * 0.4 + clin_acc * 0.6

    print(f"\n{SEP}")
    print("  SUMMARY")
    print(f"{SEP}")
    print(f"\n  {'Group':<16} {'B Accuracy':>12} {'B Coverage':>12} {'A Accuracy':>12}")
    print(f"  {'-'*16} {'-'*12} {'-'*12} {'-'*12}")
    for g, s in group_summaries.items():
        print(f"  {g:<16} {s['acc_B']:>11.1f}% {s['cov_B']:>11.1f}% {s['acc_A']:>11.1f}%")

    print(f"\n  Category 1 \u2014 Demographics (\xd70.4):  accuracy={demo_acc:.1f}%  coverage={demo_cov:.1f}%")
    print(f"  Category 2 \u2014 Clinical     (\xd70.6):  accuracy={clin_acc:.1f}%  coverage={clin_cov:.1f}%")
    print(f"\n  \u2605  OVERALL ACCURACY:  {overall:.1f}%")
    print(f"  \u2605  Missing data does NOT reduce this score \u2014 see coverage report below")

    # Error attribution
    total_err = attr_total['stage1_err'] + attr_total['stage2_err']
    print(f"\n{SEP}")
    print("  ERROR ATTRIBUTION  (accuracy failures = wrong + hallucination only)")
    print(f"{SEP}")
    if total_err > 0:
        print(f"\n  Stage 1 errors (wrong in JSON AND Excel):      "
              f"{attr_total['stage1_err']:>4}  ({attr_total['stage1_err']/total_err*100:.0f}%)")
        print(f"  Stage 2 errors (correct in JSON, wrong Excel): "
              f"{attr_total['stage2_err']:>4}  ({attr_total['stage2_err']/total_err*100:.0f}%)")
        print(f"  End-to-end correct:                            "
              f"{attr_total['e2e_correct']:>4}")
        print(f"  Stage 2 corrected a Stage 1 error (flag):      "
              f"{attr_total['stage2_fix']:>4}")

    # Coverage report
    print(f"\n{SEP}")
    print("  COVERAGE REPORT  (separate metric \u2014 missing fields do not affect accuracy)")
    print(f"{SEP}")
    print(f"\n  {'Group':<16} {'In source':>10} {'Populated':>10} {'Coverage':>10}")
    print(f"  {'-'*16} {'-'*10} {'-'*10} {'-'*10}")
    for g, s in group_summaries.items():
        b   = s['B']
        pop = b['correct'] + b['wrong'] + b['halluc']
        print(f"  {g:<16} {s['src_B']:>10} {pop:>10} {s['cov_B']:>9.1f}%")
    print()


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Final Validation — Clinical AI Pipeline')
    parser.add_argument('--docx',  required=True, help='Path to Word document (ground truth)')
    parser.add_argument('--excel', required=True, help='Path to Excel output to validate')
    parser.add_argument('--jsons', required=True, help='Directory containing case_xxx.json files')
    args = parser.parse_args()

    print(f"\nLoading files...")
    doc = docx.Document(args.docx)
    wb  = openpyxl.load_workbook(args.excel)
    ws  = wb.active

    headers    = [cell.value for cell in ws[1]]
    excel_rows = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]

    json_dir = pathlib.Path(args.jsons)
    jsons    = []
    for i in range(50):
        p = json_dir / f'case_{i:03d}.json'
        jsons.append(json.load(open(p)) if p.exists() else {})

    print(f"  Excel: {len(excel_rows)} rows")
    print(f"  JSONs: {sum(1 for j in jsons if j)} loaded")

    print(f"  Parsing ground truth from Word doc...")
    gt_patients = parse_ground_truth(doc)
    print(f"  → {len(gt_patients)} patients parsed")

    # Build field lookup
    field_lookup = {gt_key: (ecol, jpath, mtype) for gt_key, ecol, jpath, mtype in FIELDS}

    # Score all fields
    results    = {}
    attr_total = {'e2e_correct': 0, 'stage2_err': 0, 'stage1_err': 0, 'stage2_fix': 0}

    for gt_key, excel_col, json_path, match_type in FIELDS:
        r = score_field(gt_patients, excel_rows, jsons, gt_key, excel_col, json_path, match_type)
        results[gt_key] = r
        for k in attr_total:
            attr_total[k] += r['attr'][k]

    # Build group summaries
    group_summaries = {}
    for group, fields in GROUPS.items():
        gA = {'correct': 0, 'wrong': 0, 'halluc': 0, 'missing': 0}
        gB = {'correct': 0, 'wrong': 0, 'halluc': 0, 'missing': 0}
        for fk in fields:
            for k in gA:
                gA[k] += results[fk]['A'][k]
                gB[k] += results[fk]['B'][k]

        dB = gB['correct'] + gB['wrong'] + gB['halluc']
        dA = gA['correct'] + gA['wrong'] + gA['halluc']
        sB = dB + gB['missing']
        sA = dA + gA['missing']

        group_summaries[group] = {
            'acc_A': gA['correct'] / dA * 100 if dA > 0 else 0,
            'cov_A': dA / sA * 100 if sA > 0 else 0,
            'acc_B': gB['correct'] / dB * 100 if dB > 0 else 0,
            'cov_B': dB / sB * 100 if sB > 0 else 0,
            'B': gB, 'A': gA, 'src_B': sB,
        }

    print_report(results, group_summaries, attr_total)


if __name__ == '__main__':
    main()
