import os
import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.comments import Comment

# Configuration
PROJECT_ROOT = Path("/Users/joshuabhawanlall/Git Folder/Clinical_AI_Extraction_Pipeline")
INPUT_DIR = PROJECT_ROOT / "v10-Multimodal-Prose-Grid-Fusion/output/multimodal_facts"
PROTOTYPE_WORKBOOK = PROJECT_ROOT / "data/hackathon-database-prototype.xlsx"
OUTPUT_WORKBOOK = PROJECT_ROOT / "v10-Multimodal-Prose-Grid-Fusion/output/generated-database-v16-multimodal.xlsx"

def main():
    if not os.path.exists(OUTPUT_WORKBOOK.parent): os.makedirs(OUTPUT_WORKBOOK.parent)
    
    fact_files = sorted([f for f in os.listdir(INPUT_DIR) if f.endswith(".json")])
    patient_journeys = {}
    for f in fact_files:
        with open(INPUT_DIR / f, "r") as j: data = json.load(j)
        nhs = data["nhs"]
        if nhs not in patient_journeys: patient_journeys[nhs] = []
        patient_journeys[nhs].append(data)
        
    df_template = pd.read_excel(PROTOTYPE_WORKBOOK)
    template_cols = list(df_template.columns)
    final_rows = []
    
    for nhs, events in patient_journeys.items():
        master_row = {}
        # Sort by chronological event index
        sorted_events = sorted(events, key=lambda x: x["event_index"])
        
        for i, event in enumerate(sorted_events):
            facts = event["resolved_facts"]
            
            # Map events to Clusters
            # 0 -> Baseline
            # 1 -> 2nd MRI / 6-week
            # 2 -> 12-week
            
            for col, val in facts.items():
                if not val: continue
                # Sequential Slotting
                if not master_row.get(col):
                    master_row[col] = val
                elif any(kw in col for kw in ["Baseline", "Date", "CT", "MRI", "MDT"]):
                    # Identify the next tier
                    fup_col = col.replace("Baseline CT", "2nd MRI").replace("Baseline MRI", "2nd MRI").replace("1st MDT", "MDT (after 6 week")
                    if fup_col in template_cols and not master_row.get(fup_col):
                        master_row[fup_col] = val
                    else:
                        fup_12 = col.replace("Baseline CT", "12 week MRI").replace("Baseline MRI", "12 week MRI").replace("1st MDT", "MDT (after 12 week")
                        if fup_12 in template_cols and not master_row.get(fup_12):
                            master_row[fup_12] = val
                            
            master_row['Demographics: \nNHS number(d)'] = nhs

        final_rows.append(master_row)

    # Assembly
    df_final = pd.DataFrame(final_rows)
    df_final = df_final.reindex(columns=template_cols)
    wb = load_workbook(PROTOTYPE_WORKBOOK)
    ws = wb.active
    if ws.max_row >= 2: ws.delete_rows(2, ws.max_row)
    
    print(f"v10 Multimodal Assembly for {len(final_rows)} patients...")
    for r_idx, (_, row) in enumerate(df_final.iterrows(), start=2):
        for c_idx, value in enumerate(row.tolist(), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            val_str = str(value).strip() if value and str(value).lower() != "nan" else ""
            if val_str.endswith(".0"): val_str = val_str[:-2]
            if val_str:
                cell.value = val_str
                cell.comment = Comment(f"v10 Multimodal Fusion:\nSynthesized from Paragraphs & Tables.", "Gemini CLI")
            cell.number_format = '@'
            cell.data_type = 's'
            
    wb.save(OUTPUT_WORKBOOK)
    print(f"v10 Multimodal Implementation Complete. Output: {OUTPUT_WORKBOOK}")

if __name__ == "__main__":
    main()
