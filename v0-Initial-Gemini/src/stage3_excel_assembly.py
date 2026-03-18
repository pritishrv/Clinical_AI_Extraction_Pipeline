import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Configuration
INPUT_DIR = "output/json_verified"
OUTPUT_EXCEL = "output/final_database_v3.xlsx"
PROTOTYPE_PATH = "../data/hackathon-database-prototype.xlsx"

def load_verified_json():
    """
    Load all verified JSON files and return as a list of dicts.
    """
    verified_files = sorted([f for f in os.listdir(INPUT_DIR) if f.endswith(".json")])
    data = []
    for f in verified_files:
        with open(os.path.join(INPUT_DIR, f), "r") as json_file:
            data.append(json.load(json_file))
    return data

def assemble_dataframe(verified_data):
    """
    Map JSON fields to the flat database schema.
    """
    rows = []
    for case in verified_data:
        # Simplified mapping for initial version
        row = {
            "Case Index": case["case_index"],
            "Diagnosis": ", ".join(case["mapped_fields"].get("diagnosis", [])),
            "Dates": ", ".join(case["mapped_fields"].get("dates", [])),
            "Procedures": ", ".join(case["mapped_fields"].get("procedures", [])),
            "Human Verification Required": "Review NER confidence" # Placeholder flag
        }
        rows.append(row)
    return pd.DataFrame(rows)

def write_styled_excel(df, output_path):
    """
    Write the DataFrame to a styled Excel file.
    In a complete version, we would copy the prototype schema and formatting.
    """
    print(f"Writing styled Excel to {output_path}...")
    df.to_excel(output_path, index=False)
    
    # Placeholder for openpyxl styling (font, fill, etc.)
    wb = Workbook()
    ws = wb.active
    ws.title = "Prototype V1"
    
    # Minimal styling for the initial version
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # Write headers
    for col_num, column_title in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Write data
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num).value = cell_value
            
    wb.save(output_path)
    print("Excel saved successfully.")

if __name__ == "__main__":
    # Ensure relative paths work from the script location
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    os.chdir("..")
    
    if not os.path.exists(INPUT_DIR):
        print(f"Error: {INPUT_DIR} not found. Run Stage 2 first.")
    else:
        verified_data = load_verified_json()
        df = assemble_dataframe(verified_data)
        write_styled_excel(df, OUTPUT_EXCEL)
