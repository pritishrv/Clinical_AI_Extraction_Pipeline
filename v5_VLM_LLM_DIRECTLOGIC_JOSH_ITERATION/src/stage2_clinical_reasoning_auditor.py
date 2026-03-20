import os
import json
import re
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.comments import Comment
from tqdm import tqdm

# Configuration
PROJECT_ROOT = Path("/Users/joshuabhawanlall/Git Folder/Clinical_AI_Extraction_Pipeline")
ITERATION_DIR = PROJECT_ROOT / "v5_VLM_LLM_DIRECTLOGIC_JOSH_ITERATION"
JSON_DIR = ITERATION_DIR / "output/json_raw_claude"
PROTOTYPE_PATH = PROJECT_ROOT / "data/hackathon-database-prototype.xlsx"
OUTPUT_PATH = ITERATION_DIR / "output/josh_iteration_audited_database.xlsx"

# Clinical Logic Patterns
PATTERNS = {
    "T": r"\bT([0-4][a-d]?)\b",
    "N": r"\bN([0-3][a-c]?)\b",
    "M": r"\bM([01][a-c]?)\b",
    "mrT": r"mrT([0-4][a-d]?)",
    "mrN": r"mrN([0-3][a-c]?)",
    "mrTRG": r"TRG\s*([1-5])",
    "Mets": r"metastases|mets|M1",
}

def clean_nhs(nhs_str):
    if not nhs_str: return "UNKNOWN"
    return re.sub(r"\D", "", str(nhs_str))[:10]

def audit_reasoning(event):
    facts = {}
    scores = {}
    prose = (str(event.get("clinical_details", "")) + " " + str(event.get("mdt_outcome", ""))).lower()
    for key, pat in PATTERNS.items():
        match = re.search(pat, prose, re.I)
        if match:
            val = match.group(1) if match.groups() else match.group(0)
            facts[key] = val.upper()
            scores[key] = 0.7
    if any(x in prose for x in ["lung metastases", "liver metastases", "bone metastases", "m1"]):
        facts["M"] = "M1"
        scores["M"] = 0.8
    return facts, scores

def main():
    if not os.path.exists(JSON_DIR): return
    patient_journeys = {}
    json_files = sorted([f for f in os.listdir(JSON_DIR) if f.endswith(".json")])
    for f_name in json_files:
        with open(JSON_DIR / f_name, "r") as f: data = json.load(f)
        nhs = clean_nhs(data.get("patient_details", {}).get("nhs_number"))
        if nhs not in patient_journeys: patient_journeys[nhs] = []
        patient_journeys[nhs].append(data)

    df_proto = pd.read_excel(PROTOTYPE_PATH)
    target_cols = list(df_proto.columns)
    final_rows = []
    audit_highlights = []

    print(f"Auditing {len(patient_journeys)} clinical journeys...")
    for p_idx, (nhs, events) in enumerate(patient_journeys.items()):
        events = sorted(events, key=lambda x: x.get("mdt_meeting_date", ""))
        master_row = {col: None for col in target_cols}
        first = events[0]
        master_row['Demographics: \nNHS number(d)'] = nhs
        master_row['Demographics: \nDOB(a)'] = first.get("patient_details", {}).get("dob")
        master_row['Demographics: \nGender(e)'] = first.get("patient_details", {}).get("gender", "").split("(")[0].strip()
        
        for i, event in enumerate(events):
            inferred_facts, inferred_scores = audit_reasoning(event)
            prefix = "Baseline" if i == 0 else "2nd MRI" if i == 1 else "12 week MRI"
            if "T" in inferred_facts:
                col = f"Baseline CT: T(h)" if i == 0 else f"{prefix}: mrT"
                if col in master_row:
                    master_row[col] = inferred_facts["T"]
                    if inferred_scores.get("T", 1.0) < 1.0:
                        audit_highlights.append((p_idx + 2, target_cols.index(col) + 1, "Inferred from MDT prose sequence."))
            if i == 0: master_row['1st MDT: date(i)'] = event.get("mdt_meeting_date")
            elif i == 1: master_row['MDT (after 6 week: Date'] = event.get("mdt_meeting_date")
        final_rows.append(master_row)

    df_final = pd.DataFrame(final_rows)
    df_final.to_excel(OUTPUT_PATH, index=False)
    wb = load_workbook(OUTPUT_PATH)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for r, c, msg in audit_highlights:
        cell = ws.cell(row=r, column=c)
        cell.fill = yellow_fill
        cell.comment = Comment(f"CLINICAL AUDIT: {msg}", "NHS Consultant Auditor")
    wb.save(OUTPUT_PATH)
    print(f"v5-Josh Iteration Complete. Database: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
