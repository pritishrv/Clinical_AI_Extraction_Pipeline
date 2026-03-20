import os
import json
import re
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from tqdm import tqdm
import anthropic
import time

# Configuration
PROJECT_ROOT = Path("/Users/joshuabhawanlall/Git Folder/Clinical_AI_Extraction_Pipeline")
ITERATION_DIR = PROJECT_ROOT / "v5_VLM_LLM_DIRECTLOGIC_JOSH_ITERATION"
JSON_DIR = ITERATION_DIR / "output/json_raw_claude"
PROTOTYPE_PATH = PROJECT_ROOT / "data/hackathon-database-prototype.xlsx"
OUTPUT_PATH = ITERATION_DIR / "output/josh_iteration_audited_database.xlsx"

# Import Private Key
try:
    import config_private
    client = anthropic.Anthropic(api_key=config_private.ANTHROPIC_API_KEY)
except Exception as e:
    print(f"Error: Could not load config_private.py: {e}")
    client = None

def get_chapters(all_cols):
    """Splits 88 columns into 3 manageable Clinical Chapters."""
    c1 = all_cols[:30] # Presentation & Baseline
    c2 = all_cols[30:60] # Treatment & Response
    c3 = all_cols[60:] # Surgery & Surveillance
    return [c1, c2, c3]

def call_consultant_chapter(journey_summary, chapter_cols, chapter_num):
    """Calls Claude for a specific subset of the schema."""
    if not client: return None
    
    prompt = f"""
    You are a Senior NHS Colorectal Consultant. Reason over this patient's journey for CHAPTER {chapter_num}.
    
    PATIENT DATA:
    {journey_summary}

    TARGET COLUMNS FOR THIS CHAPTER:
    {json.dumps(chapter_cols)}

    RULES:
    1. Reason as a doctor. Infer stages/dates from prose if not explicit.
    2. CHRONOLOGY: Ensure sequential mapping (e.g. 1st MDT vs 2nd MDT).
    3. Return ONLY a JSON object where each key is a column name and value is: {{"value": "val", "confidence": 0.0-1.0, "rationale": "text"}}.
    4. If no data for a column, set value to null.
    """

    for attempt in range(3):
        try:
            response = client.messages.create(
                model="claude-3-haiku-20240307",
                max_tokens=4096,
                temperature=0,
                messages=[{"role": "user", "content": prompt}]
            )
            content = response.content[0].text
            json_match = re.search(r'\{.*\}', content, re.DOTALL)
            if json_match:
                return json.loads(json_match.group(0))
        except Exception as e:
            print(f"Chapter {chapter_num} Error (Attempt {attempt}): {e}")
            time.sleep(2)
    return {}

def main():
    if not os.path.exists(JSON_DIR): return
    
    # 1. Group by Patient
    patient_journeys = {}
    json_files = sorted([f for f in os.listdir(JSON_DIR) if f.endswith(".json")])
    for f_name in json_files:
        with open(JSON_DIR / f_name, "r") as f: data = json.load(f)
        nhs = re.sub(r"\D", "", str(data.get("patient_details", {}).get("nhs_number", "UNKNOWN")))[:10]
        if nhs not in patient_journeys: patient_journeys[nhs] = []
        patient_journeys[nhs].append(data)

    df_proto = pd.read_excel(PROTOTYPE_PATH)
    target_cols = list(df_proto.columns)
    chapters = get_chapters(target_cols)
    final_rows = []
    audit_data = []

    print(f"NHS Consultant: Multi-Pass Chapter Audit for {len(patient_journeys)} patients...")
    for p_idx, (nhs, events) in enumerate(tqdm(patient_journeys.items())):
        events = sorted(events, key=lambda x: x.get("mdt_meeting_date", ""))
        journey_summary = ""
        for i, ev in enumerate(events):
            journey_summary += f"\n[MDT {i} | {ev.get('mdt_meeting_date')}]\n"
            journey_summary += f"Staging: {ev.get('staging_and_diagnosis')}\n"
            journey_summary += f"Outcome: {ev.get('mdt_outcome')}\n"
            journey_summary += f"Clinical: {ev.get('clinical_details')}\n"

        master_row = {col: None for col in target_cols}
        master_row['Demographics: \nNHS number(d)'] = nhs
        
        # Reason through each chapter sequentially
        for c_idx, col_subset in enumerate(chapters, 1):
            chapter_results = call_consultant_chapter(journey_summary, col_subset, c_idx)
            if chapter_results:
                for col, info in chapter_results.items():
                    if col in master_row and info.get("value"):
                        master_row[col] = info["value"]
                        if info.get("confidence", 1.0) < 1.0:
                            audit_data.append((len(final_rows) + 2, target_cols.index(col) + 1, info.get("rationale", "Inferred")))
            time.sleep(0.5)

        final_rows.append(master_row)

    # 3. Assemble and Styles
    df_final = pd.DataFrame(final_rows)
    df_final.to_excel(OUTPUT_PATH, index=False)
    wb = load_workbook(OUTPUT_PATH)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for r, c, msg in audit_data:
        cell = ws.cell(row=r, column=c)
        cell.fill = yellow_fill
        cell.comment = Comment(msg, "NHS Consultant Auditor")
    wb.save(OUTPUT_PATH)
    print(f"\nAudit Session Complete. Master Journey Database: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
