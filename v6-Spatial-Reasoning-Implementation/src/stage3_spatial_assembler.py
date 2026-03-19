import os
import json
import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.comments import Comment
from datetime import datetime

# Configuration
PROJECT_ROOT = Path("/Users/joshuabhawanlall/Git Folder/Clinical_AI_Extraction_Pipeline")
INPUT_DIR = PROJECT_ROOT / "v6-Spatial-Reasoning-Implementation/output/spatial_facts"
PROTOTYPE_WORKBOOK = PROJECT_ROOT / "data/hackathon-database-prototype.xlsx"
OUTPUT_WORKBOOK = PROJECT_ROOT / "v6-Spatial-Reasoning-Implementation/output/generated-database-v6-spatial.xlsx"

def parse_date(date_str):
    if not date_str: return datetime.min
    try: return datetime.strptime(str(date_str), "%d/%m/%Y")
    except: return datetime.min

def main():
    if not os.path.exists(OUTPUT_WORKBOOK.parent): os.makedirs(OUTPUT_WORKBOOK.parent)
    
    fact_files = sorted([f for f in os.listdir(INPUT_DIR) if f.endswith(".json")])
    journey_store = {}
    
    for f in fact_files:
        with open(INPUT_DIR / f, "r") as j: data = json.load(j)
        nhs = data["patient_nhs"]
        if nhs not in journey_store:
            journey_store[nhs] = {"events": []}
        journey_store[nhs]["events"].append(data)
        
    df_template = pd.read_excel(PROTOTYPE_WORKBOOK)
    template_cols = list(df_template.columns)
    final_rows = []
    
    for nhs, patient in journey_store.items():
        master_row = {}
        # Chronological sort of events for this patient
        sorted_events = sorted(patient["events"], key=lambda x: parse_date(x["spatial_facts"].get("mdt_date")))
        
        for i, event in enumerate(sorted_events):
            facts = event["spatial_facts"]
            
            # 1. ADDITIVE MAPPING: Ingest all discovered facts
            for col, val in facts.items():
                if not val: continue
                # If slot empty, fill it
                if not master_row.get(col):
                    master_row[col] = val
                # Sequential DRIFT: Route subsequent staging to follow-up columns
                elif "Baseline" in col:
                    fup_col = col.replace("Baseline CT", "2nd MRI").replace("Baseline MRI", "2nd MRI")
                    if fup_col in template_cols and not master_row.get(fup_col):
                        master_row[fup_col] = val
                    else:
                        fup_12 = col.replace("Baseline CT", "12 week MRI").replace("Baseline MRI", "12 week MRI")
                        if fup_12 in template_cols and not master_row.get(fup_12):
                            master_row[fup_12] = val
                            
            # 2. Static Demographics (Safety Guard)
            master_row['Demographics: \nNHS number(d)'] = nhs
            if i == 0:
                master_row['1st MDT: date(i)'] = facts.get("mdt_date")
            elif i == 1:
                master_row['MDT (after 6 week: Date'] = facts.get("mdt_date")
            elif i == 2:
                master_row['MDT (after 12 week): Date'] = facts.get("mdt_date")

        final_rows.append(master_row)

    # Final Generation
    df_final = pd.DataFrame(final_rows)
    df_final = df_final.reindex(columns=template_cols)
    wb = load_workbook(PROTOTYPE_WORKBOOK)
    ws = wb.active
    if ws.max_row >= 2: ws.delete_rows(2, ws.max_row)
    
    print("v6 Final Cumulative Spatial Assembly...")
    for r_idx, (_, row) in enumerate(df_final.iterrows(), start=2):
        for c_idx, value in enumerate(row.tolist(), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            val_str = str(value).strip() if value and str(value).lower() != "nan" else ""
            if val_str.endswith(".0"): val_str = val_str[:-2]
            if val_str:
                cell.value = val_str
                cell.comment = Comment(f"v6 Spatial Reasoning:\nResolved from structural grid coordinates.", "Gemini CLI")
            cell.number_format = '@'
            cell.data_type = 's'
            
    wb.save(OUTPUT_WORKBOOK)
    print(f"v6 Spatial Fusion Complete. Output: {OUTPUT_WORKBOOK}")

if __name__ == "__main__":
    main()
