from copy import copy
from pathlib import Path

from openpyxl import load_workbook


def write_styled_workbook(dataframe, template_path, output_path):
    template_path = Path(template_path)
    output_path = Path(output_path)

    workbook = load_workbook(template_path)
    worksheet = workbook[workbook.sheetnames[0]]

    template_header_row = 1
    template_data_row = 2

    template_headers = [
        worksheet.cell(template_header_row, column_index).value
        for column_index in range(1, len(dataframe.columns) + 1)
    ]
    template_row_height = worksheet.row_dimensions[template_data_row].height
    template_data_styles = [
        copy(worksheet.cell(template_data_row, column_index)._style)
        for column_index in range(1, len(dataframe.columns) + 1)
    ]
    template_fonts = [
        copy(worksheet.cell(template_data_row, column_index).font)
        for column_index in range(1, len(dataframe.columns) + 1)
    ]
    template_fills = [
        copy(worksheet.cell(template_data_row, column_index).fill)
        for column_index in range(1, len(dataframe.columns) + 1)
    ]
    template_borders = [
        copy(worksheet.cell(template_data_row, column_index).border)
        for column_index in range(1, len(dataframe.columns) + 1)
    ]
    template_alignments = [
        copy(worksheet.cell(template_data_row, column_index).alignment)
        for column_index in range(1, len(dataframe.columns) + 1)
    ]
    template_number_formats = [
        worksheet.cell(template_data_row, column_index).number_format
        for column_index in range(1, len(dataframe.columns) + 1)
    ]
    template_protections = [
        copy(worksheet.cell(template_data_row, column_index).protection)
        for column_index in range(1, len(dataframe.columns) + 1)
    ]

    # Keep the template sheet name and workbook-level presentation.
    for column_index, header in enumerate(template_headers, start=1):
        worksheet.cell(template_header_row, column_index).value = header

    # Remove any existing example data rows before writing generated rows.
    if worksheet.max_row >= template_data_row:
        worksheet.delete_rows(template_data_row, worksheet.max_row - template_data_row + 1)

    # Recreate generated rows using the template's sample data-row style.
    worksheet.insert_rows(template_data_row, amount=len(dataframe))

    for row_offset, (_, row) in enumerate(dataframe.iterrows(), start=0):
        target_row_index = template_data_row + row_offset

        if template_row_height is not None:
            worksheet.row_dimensions[target_row_index].height = template_row_height

        for column_index, value in enumerate(row.tolist(), start=1):
            target_cell = worksheet.cell(target_row_index, column_index)
            target_cell._style = copy(template_data_styles[column_index - 1])
            target_cell.font = copy(template_fonts[column_index - 1])
            target_cell.fill = copy(template_fills[column_index - 1])
            target_cell.border = copy(template_borders[column_index - 1])
            target_cell.alignment = copy(template_alignments[column_index - 1])
            target_cell.number_format = template_number_formats[column_index - 1]
            target_cell.protection = copy(template_protections[column_index - 1])
            target_cell.value = None if value == "" else value

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
