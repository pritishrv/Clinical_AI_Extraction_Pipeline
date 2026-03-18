import os
import json
import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook

# Configuration
PROJECT_ROOT = Path("/Users/joshuabhawanlall/Git Folder/Clinical_AI_Extraction_Pipeline")
INPUT_DIR = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/master_harvest"
PROTOTYPE_WORKBOOK = PROJECT_ROOT / "data/hackathon-database-prototype.xlsx"
OUTPUT_WORKBOOK = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/generated-database-v10-master.xlsx"

def clean_val(v):
    return str(v).strip().replace("\n", " ")

def extract_facts(text):
    facts = {}
    # Basics
    nhs = re.search(r"NHS Number:\s*(\d+)", text, re.I)
    mrn = re.search(r"Hospital Number:\s*(\d+)", text, re.I)
    dob = re.search(r"(\d{1,2}/\d{1,2}/\d{4})\(a\)", text)
    if nhs: facts["Demographics: \nNHS number(d)"] = nhs.group(1)
    if mrn: facts["Demographics: MRN(c)"] = mrn.group(1)
    if dob: facts["Demographics: \nDOB(a)"] = dob.group(1)
    
    # Path
    if "Diagnosis:" in text: facts["Histology: Biopsy result(g)"] = text.split("Diagnosis:")[1].split("|")[0].strip()
    
    # Staging
    ts = re.findall(r"\bT(\d[a-d]?)\b", text)
    ns = re.findall(r"\bN(\d[a-c]?)\b", text)
    mts = re.findall(r"mrT(\d[a-d]?)\b", text)
    mns = re.findall(r"mrN(\d[a-c]?)\b", text)
    if ts: facts["Baseline CT: T(h)"] = ts[0]
    if ns: facts["Baseline CT: N(h)"] = ns[0]
    if mts: facts["Baseline MRI: mrT(h)"] = mts[0]
    if mns: facts["Baseline MRI: mrN(h)"] = mns[0]
    
    # Treatment
    drugs = ["FOLFOX", "CAPOX", "capecitabine", "Oxaliplatin"]
    found = [d for d in drugs if d in text]
    if found: facts["Chemotherapy: Drugs"] = ", ".join(found)
    
    # Dates
    dates = re.findall(r"(\d{1,2}/\d{1,2}/\d{4})", text)
    facts["dates"] = dates
    
    return facts

def main():
    if not os.path.exists(OUTPUT_WORKBOOK.parent): os.makedirs(OUTPUT_WORKBOOK.parent)
    files = sorted([f for f in os.listdir(INPUT_DIR) if f.endswith(".json")])
    
    # 1. Group by Patient
    all_patients = {}
    for f in files:
        with open(INPUT_DIR / f) as j: case = json.load(j)
        facts = extract_facts(case["text"])
        nhs = facts.get("Demographics: \nNHS number(d)", f"CASE_{case['case_index']}")
        if nhs not in all_patients: all_patients[nhs] = []
        all_patients[nhs].append(facts)
        
    # 2. Master Timeline Slotting
    final_rows = []
    for nhs, journey in all_patients.items():
        master = {}
        # Chronological sequence of documents
        for i, doc in enumerate(journey):
            if i == 0:
                # First document is Baseline
                master.update(doc)
            else:
                # Sequential Slotting: Route data to follow-up columns
                for k, v in doc.items():
                    if k == "dates" or not v: continue
                    if "Baseline" in k:
                        # Drift to 2nd MRI slots
                        fup_k = k.replace("Baseline CT", "2nd MRI").replace("Baseline MRI", "2nd MRI")
                        if fup_k not in master: master[fup_k] = v
                    elif not master.get(k):
                        master[k] = v
            
            # Slot dates into sequence
            for idx, d in enumerate(doc.get("dates", [])):
                if idx == 0 and not master.get("Baseline CT: Date(h)"): master["Baseline CT: Date(h)"] = d
                elif idx == 1 and not master.get("Baseline MRI: date(h)"): master["Baseline MRI: date(h)"] = d
                elif idx == 2 and not master.get("2nd MRI: Date"): master["2nd MRI: Date"] = d
                elif idx == 3 and not master.get("12 week MRI: Date"): master["12 week MRI: Date"] = d

        final_rows.append(master)

    # 3. Assemble Excel
    df = pd.DataFrame(final_rows)
    template_cols = list(pd.read_excel(PROTOTYPE_WORKBOOK).columns)
    df = df.reindex(columns=template_cols)
    
    wb = load_workbook(PROTOTYPE_WORKBOOK)
    ws = wb.active
    if ws.max_row >= 2: ws.delete_rows(2, ws.max_row)
    
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for c_idx, value in enumerate(row.tolist(), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            val = str(value).strip() if value and str(value).lower() != "nan" else ""
            if val.endswith(".0"): val = val[:-2]
            cell.value = val if val else None
            cell.number_format = '@'
            cell.data_type = 's'
            
    wb.save(OUTPUT_WORKBOOK)
    print(f"Master Assembly Complete. Output: {OUTPUT_WORKBOOK}")

if __name__ == "__main__":
    main()
