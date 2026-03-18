import os
import json
import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook

# Configuration
PROJECT_ROOT = Path("/Users/joshuabhawanlall/Git Folder/Clinical_AI_Extraction_Pipeline")
ZONAL_DIR = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/json_zonal"
MAPPED_DIR = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/mapped_harvest"
PROTOTYPE_WORKBOOK = PROJECT_ROOT / "data/hackathon-database-prototype.xlsx"
OUTPUT_WORKBOOK = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/generated-database-v6-harvester.xlsx"

def map_zonal_basics(zonal_data):
    """
    Extracts core precise IDs from zonal data to ensure integrity.
    """
    row = {}
    ref = zonal_data.get("REFERRAL", "")
    base = zonal_data.get("BASELINE", "")
    
    nhs = re.search(r"NHS Number:\s*(\d+)", ref, re.I)
    mrn = re.search(r"Hospital Number:\s*(\d+)", ref, re.I)
    dob = re.search(r"(\d{1,2}/\d{1,2}/\d{4})\(a\)", ref)
    
    if nhs: row["Demographics: \nNHS number(d)"] = nhs.group(1)
    if mrn: row["Demographics: MRN(c)"] = mrn.group(1)
    if dob: row["Demographics: \nDOB(a)"] = dob.group(1)
    
    # Initials
    name_match = re.search(r"([A-Z\s']+)\(b\)", ref)
    if name_match:
        pts = name_match.group(1).strip().split()
        if len(pts) >= 2: row["Demographics: Initials(b)"] = f"{pts[0][0].upper()}{pts[-1][0].upper()}"

    if "Male" in ref: row["Demographics: \nGender(e)"] = "Male"
    elif "Female" in ref: row["Demographics: \nGender(e)"] = "Female"
    
    return row

def main():
    if not os.path.exists(OUTPUT_WORKBOOK.parent):
        os.makedirs(OUTPUT_WORKBOOK.parent)
        
    zonal_files = sorted([f for f in os.listdir(ZONAL_DIR) if f.endswith(".json")])
    
    final_rows = []
    for f in zonal_files:
        # 1. Precise Zonal Basics
        with open(ZONAL_DIR / f) as j:
            zonal_data = json.load(j)
        precise_row = map_zonal_basics(zonal_data)
            
        # 2. High Density Semantic Mapped data
        mapped_file = MAPPED_DIR / f.replace("_zonal.json", "_mapped.json")
        mapped_cols = {}
        if os.path.exists(mapped_file):
            with open(mapped_file) as j:
                mapped_cols = json.load(j).get("mapped_columns", {})
        
        # 3. MERGE (Precise IDs overwrite fuzzy mapping)
        combined_row = {**mapped_cols, **precise_row}
        final_rows.append(combined_row)
        
    df = pd.DataFrame(final_rows)
    template_cols = list(pd.read_excel(PROTOTYPE_WORKBOOK).columns)
    df = df.reindex(columns=template_cols)
    
    # Write styled output
    wb = load_workbook(PROTOTYPE_WORKBOOK)
    ws = wb.active
    start_row = 2
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row)

    for r_idx, (_, row) in enumerate(df.iterrows(), start=start_row):
        for c_idx, value in enumerate(row.tolist(), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            val = str(value).strip() if value and str(value).lower() != "nan" else ""
            if val.endswith(".0"): val = val[:-2]
            cell.value = val if val else None
            cell.number_format = '@'
            cell.data_type = 's'
            
    wb.save(OUTPUT_WORKBOOK)
    print(f"Assembly Complete. Output: {OUTPUT_WORKBOOK}")

if __name__ == "__main__":
    main()
