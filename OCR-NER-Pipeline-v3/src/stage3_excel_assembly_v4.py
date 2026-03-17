import os
import json
import pandas as pd
import re
from pathlib import Path
from sys import path
from copy import copy
from openpyxl import load_workbook

# Configuration
PROJECT_ROOT = Path("/Users/joshuabhawanlall/Git Folder/Clinical_AI_Extraction_Pipeline")
INPUT_DIR = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/json_ner_v4"
PROTOTYPE_WORKBOOK = PROJECT_ROOT / "data/hackathon-database-prototype.xlsx"
OUTPUT_WORKBOOK = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/generated-database-v4.xlsx"

def map_case_to_row(case_data):
    full_text = case_data["full_text"]
    entities = case_data["entities"]
    
    data = {}
    
    # --- 1. Core Identifiers ---
    nhs = re.search(r"NHS Number:\s*(\d+)", full_text, re.I)
    mrn = re.search(r"Hospital Number:\s*(\d+)", full_text, re.I)
    dob = re.search(r"(\d{1,2}/\d{1,2}/\d{4})\(a\)", full_text)
    if nhs: data["Demographics: \nNHS number(d)"] = nhs.group(1)
    if mrn: data["Demographics: MRN(c)"] = mrn.group(1)
    if dob: data["Demographics: \nDOB(a)"] = dob.group(1)
    
    name_match = re.search(r"([A-Z\s']+)\(b\)", full_text)
    if name_match:
        pts = name_match.group(1).strip().split()
        if len(pts) >= 2: data["Demographics: Initials(b)"] = f"{pts[0][0].upper()}{pts[-1][0].upper()}"

    if "Male" in full_text[:1000]: data["Demographics: \nGender(e)"] = "Male"
    elif "Female" in full_text[:1000]: data["Demographics: \nGender(e)"] = "Female"

    # --- 2. Pathology (Ultra Sweep) ---
    if "Diagnosis:" in full_text:
        diag = full_text.split("Diagnosis:")[1].split("|")[0].strip().replace("\n", " ")
        data["Histology: Biopsy result(g)"] = diag
        if "rectum" in diag.lower(): data["Demographics: \nState site of previous cancer(f)"] = "Rectum"
        elif "colon" in diag.lower(): data["Demographics: \nState site of previous cancer(f)"] = "Colon"

    for ent in entities:
        if ent["negated"]: continue
        l, t = ent["label"], ent["text"]
        if l == "DUKES": data["Dukes:"] = t.replace("Dukes", "").strip()
        if l == "DIFFERENTIATION": data["Histology: Biopsy result(g)"] = t
        if l == "MMR_STATUS": data["Histology: \nMMR status(g/h)"] = "Deficient" if "deficient" in t.lower() else "Proficient"

    # --- 3. Imaging & Staging (Deep Mining) ---
    data["Baseline CT: M(h)"] = "1" if "metastases" in full_text.lower() and "no metastases" not in full_text.lower() else "0"
    
    ts = re.findall(r"\bT(\d[a-d]?)\b", full_text)
    ns = re.findall(r"\bN(\d[a-c]?)\b", full_text)
    mts = re.findall(r"mrT(\d[a-d]?)\b", full_text)
    mns = re.findall(r"mrN(\d[a-c]?)\b", full_text)
    trgs = re.findall(r"mrTRG\s*([1-5])", full_text)
    
    if ts: data["Baseline CT: T(h)"] = ts[0]
    if ns: data["Baseline CT: N(h)"] = ns[0]
    if mts: 
        data["Baseline MRI: mrT(h)"] = mts[0]
        if len(mts) > 1: data["2nd MRI: mrT"] = mts[1]
        if len(mts) > 2: data["12 week MRI: mrT"] = mts[2]
    if mns: 
        data["Baseline MRI: mrN(h)"] = mns[0]
        if len(mns) > 1: data["2nd MRI: mrN"] = mns[1]
        if len(mns) > 2: data["12 week MRI: mrN"] = mns[2]
    if trgs:
        data["2nd MRI: mrTRG score "] = trgs[0]
        if len(trgs) > 1: data["12 week MRI: mrTRG score "] = trgs[1]

    if "emvi positive" in full_text.lower() or "emvi +" in full_text.lower(): data["Baseline MRI: mrEMVI(h)"] = "Positive"
    if "crm clear" in full_text.lower(): data["Baseline MRI: mrCRM(h)"] = "Clear"

    # --- 4. Treatment (Fill Gap strategy) ---
    for ent in entities:
        if ent["negated"]: continue
        l, t = ent["label"], ent["text"]
        if l == "CHEMO_DRUG": data["Chemotherapy: Drugs"] = t
        if l == "CHEMO_CYCLE": data["Chemotherapy: Cycles"] = t
        if l == "RT_DOSE": data["Radiotheapy: Total dose"] = t
        if l == "RT_FRACTIONS": data["Radiotheapy: Boost"] = t
        if l == "INTENT": 
            data["Chemotherapy: Treatment goals  \n(curative, palliative)"] = t
            data["Surgery: Intent, pre-neoadjuvant therapy"] = t
        if l == "SURGERY": data["Surgery: Intent, pre-neoadjuvant therapy"] = t
        if l == "W_AND_W": data["Watch and wait: Entered watch + wait, date of MDT ?"] = "Yes"

    # --- 5. Dates (Deep Sequence Miner) ---
    all_dates = re.findall(r"(\d{1,2}/\d{1,2}/\d{4})", full_text)
    if all_dates:
        data["1st MDT: date(i)"] = all_dates[-1]
        if len(all_dates) > 1: data["Baseline CT: Date(h)"] = all_dates[0]
        if len(all_dates) > 2: data["Baseline MRI: date(h)"] = all_dates[1]
        if len(all_dates) > 3: data["2nd MRI: Date"] = all_dates[2]
        if len(all_dates) > 4: data["12 week MRI: Date"] = all_dates[3]
        if len(all_dates) > 5: data["Radiotherapy: Dates"] = all_dates[4]
        if len(all_dates) > 6: data["Surgery: Date of surgery "] = all_dates[5]

    # --- 6. Outcomes & CEA ---
    if "Outcome:" in full_text:
        out = full_text.split("Outcome:")[1].split("||")[0].strip().replace("\n", " ")
        data["MDT (after 6 week: Decision "] = out
        if any(x in out.lower() for x in ["surgery", "resection"]): data["1st MDT: Treatment approach \n(TNT, downstaging chemotherapy, downstaging nCRT, downstaging shortcourse RT, Papillon +/- EBRT, straight to surgery(h)"] = "straight to surgery"

    cea = re.search(r"CEA\s*[:\-\u2013]?\s*([\d\.]+)", full_text, re.I)
    if cea: data["CEA: Value"] = cea.group(1)

    return data

def write_gold_standard(dataframe, template_path, output_path):
    wb = load_workbook(template_path)
    ws = wb.active
    start_row = 2
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row)

    for r_idx, (_, row) in enumerate(dataframe.iterrows(), start=start_row):
        for c_idx, value in enumerate(row.tolist(), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            val = str(value).strip() if value and str(value).lower() != "nan" else ""
            if val.endswith(".0"): val = val[:-2]
            cell.value = val if val else None
            cell.number_format = '@'
            cell.data_type = 's'
    wb.save(output_path)

def main():
    if not os.path.exists(OUTPUT_WORKBOOK.parent): os.makedirs(OUTPUT_WORKBOOK.parent)
    raw_files = sorted([f for f in os.listdir(INPUT_DIR) if f.endswith(".json")])
    all_data = []
    for f in raw_files:
        with open(INPUT_DIR / f) as j:
            all_data.append(map_case_to_row(json.load(j)))
    df = pd.DataFrame(all_data)
    template_cols = list(pd.read_excel(PROTOTYPE_WORKBOOK).columns)
    df = df.reindex(columns=template_cols)
    write_gold_standard(df, PROTOTYPE_WORKBOOK, OUTPUT_WORKBOOK)
    print(f"Generated Gold Standard database saved to {OUTPUT_WORKBOOK}")

if __name__ == "__main__":
    main()
