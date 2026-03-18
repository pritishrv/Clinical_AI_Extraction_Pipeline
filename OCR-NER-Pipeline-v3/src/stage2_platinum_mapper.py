import os
import json
import re
import pandas as pd
from pathlib import Path
from sys import path
from copy import copy
from openpyxl import load_workbook

# Configuration
PROJECT_ROOT = Path("/Users/joshuabhawanlall/Git Folder/Clinical_AI_Extraction_Pipeline")
INPUT_DIR = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/json_zonal"
PROTOTYPE_WORKBOOK = PROJECT_ROOT / "data/hackathon-database-prototype.xlsx"
OUTPUT_WORKBOOK = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/generated-database-platinum.xlsx"

def map_zonal_to_row(zonal_data):
    row_data = {}
    
    ref = zonal_data["REFERRAL"]
    base = zonal_data["BASELINE"]
    treat = zonal_data["PRIMARY_TREATMENT"]
    fup = zonal_data["FOLLOW_UP"]
    full_text = f"{ref} || {base} || {treat} || {fup}"
    
    # --- 1. IDs ---
    nhs = re.search(r"NHS Number:\s*(\d+)", ref, re.I)
    mrn = re.search(r"Hospital Number:\s*(\d+)", ref, re.I)
    dob = re.search(r"(\d{1,2}/\d{1,2}/\d{4})\(a\)", ref)
    if nhs: row_data["Demographics: \nNHS number(d)"] = nhs.group(1)
    if mrn: row_data["Demographics: MRN(c)"] = mrn.group(1)
    if dob: row_data["Demographics: \nDOB(a)"] = dob.group(1)
    
    name_match = re.search(r"([A-Z\s']+)\(b\)", ref)
    if name_match:
        pts = name_match.group(1).strip().split()
        if len(pts) >= 2: row_data["Demographics: Initials(b)"] = f"{pts[0][0].upper()}{pts[-1][0].upper()}"

    if "Male" in ref[:1000]: row_data["Demographics: \nGender(e)"] = "Male"
    elif "Female" in ref[:1000]: row_data["Demographics: \nGender(e)"] = "Female"

    # --- 2. Pathology & Histo ---
    if "Diagnosis:" in base:
        diag = base.split("Diagnosis:")[1].split("|")[0].strip().replace("\n", " ")
        row_data["Histology: Biopsy result(g)"] = diag
        if "rectum" in diag.lower(): row_data["Demographics: \nState site of previous cancer(f)"] = "Rectum"
        elif "colon" in diag.lower(): row_data["Demographics: \nState site of previous cancer(f)"] = "Colon"
    
    if "MMR deficient" in full_text: row_data["Histology: \nMMR status(g/h)"] = "Deficient"
    elif "MMR proficient" in full_text: row_data["Histology: \nMMR status(g/h)"] = "Proficient"

    # --- 3. Staging (Ultra Density Mapping) ---
    ts = re.findall(r"\bT(\d[a-d]?)\b", full_text)
    ns = re.findall(r"\bN(\d[a-c]?)\b", full_text)
    mts = re.findall(r"mrT(\d[a-d]?)\b", full_text)
    mns = re.findall(r"mrN(\d[a-c]?)\b", full_text)
    trgs = re.findall(r"TRG\s*([1-5])", full_text, re.I)
    
    if ts: row_data["Baseline CT: T(h)"] = ts[0]
    if ns: row_data["Baseline CT: N(h)"] = ns[0]
    if mts: 
        row_data["Baseline MRI: mrT(h)"] = mts[0]
        if len(mts) > 1: row_data["2nd MRI: mrT"] = mts[1]
        if len(mts) > 2: row_data["12 week MRI: mrT"] = mts[2]
    if mns: 
        row_data["Baseline MRI: mrN(h)"] = mns[0]
        if len(mns) > 1: row_data["2nd MRI: mrN"] = mns[1]
        if len(mns) > 2: row_data["12 week MRI: mrN"] = mns[2]
    if trgs:
        row_data["2nd MRI: mrTRG score "] = trgs[0]
        if len(trgs) > 1: row_data["12 week MRI: mrTRG score "] = trgs[1]

    row_data["Baseline CT: M(h)"] = "1" if "metastases" in full_text.lower() and "no metastases" not in full_text.lower() else "0"
    if "emvi positive" in full_text.lower() or "emvi +" in full_text.lower(): row_data["Baseline MRI: mrEMVI(h)"] = "Positive"
    if "crm clear" in full_text.lower(): row_data["Baseline MRI: mrCRM(h)"] = "Clear"

    # --- 4. Treatment & Intent ---
    drugs = ["FOLFOX", "CAPOX", "capecitabine", "5-FU", "Oxaliplatin"]
    found_drugs = [d for d in drugs if d in full_text]
    if found_drugs: 
        row_data["Chemotherapy: Drugs"] = ", ".join(set(found_drugs))
        row_data["Chemotherapy: Treatment goals  \n(curative, palliative)"] = "Curative"
    
    rt_dose = re.search(r"(\d+\s*Gy)", full_text)
    if rt_dose: row_data["Radiotheapy: Total dose"] = rt_dose.group(1)
    
    if "surgery" in treat.lower() or "resection" in treat.lower() or "hemicolectomy" in treat.lower():
        row_data["Surgery: Intent, pre-neoadjuvant therapy"] = "Resection"
    if "watch and wait" in full_text.lower():
        row_data["Watch and wait: Entered watch + wait, date of MDT ?"] = "Yes"

    # --- 5. Dates (Deep Miner) ---
    all_dates = re.findall(r"(\d{1,2}/\d{1,2}/\d{4})", full_text)
    if all_dates:
        row_data["1st MDT: date(i)"] = all_dates[-1]
        if len(all_dates) > 1: row_data["Baseline CT: Date(h)"] = all_dates[0]
        if len(all_dates) > 2: row_data["Baseline MRI: date(h)"] = all_dates[1]
        if len(all_dates) > 3: row_data["2nd MRI: Date"] = all_dates[2]
        if len(all_dates) > 4: row_data["12 week MRI: Date"] = all_dates[3]
        if len(all_dates) > 5: row_data["Radiotherapy: Dates"] = all_dates[4]

    # --- 6. Prose Outcomes ---
    if "Outcome:" in full_text:
        row_data["MDT (after 6 week: Decision "] = full_text.split("Outcome:")[1].split("||")[0].strip().replace("\n", " ")

    cea = re.search(r"CEA\s*[:\-\u2013]?\s*([\d\.]+)", full_text, re.I)
    if cea: row_data["CEA: Value"] = cea.group(1)

    return row_data

def write_platinum_styled(dataframe, template_path, output_path):
    wb = load_workbook(template_path)
    ws = wb.active
    start_row = 2
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row)

    for r_idx, (_, row) in enumerate(dataframe.iterrows(), start=start_row):
        for c_idx, value in enumerate(row.tolist(), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            val = str(value).strip() if value and str(value).lower() != "nan" else ""
            if val.endswith(".0"): val = val[:-2]
            cell.value = val if val else None
            cell.number_format = '@' 
            cell.data_type = 's'
    wb.save(output_path)

def main():
    if not os.path.exists(OUTPUT_WORKBOOK.parent): os.makedirs(OUTPUT_WORKBOOK.parent)
    zonal_files = sorted([f for f in os.listdir(INPUT_DIR) if f.endswith(".json")])
    all_rows = [map_zonal_to_row(json.load(open(INPUT_DIR / f))) for f in zonal_files]
    df = pd.DataFrame(all_rows)
    template_cols = list(pd.read_excel(PROTOTYPE_WORKBOOK).columns)
    df = df.reindex(columns=template_cols)
    write_platinum_styled(df, PROTOTYPE_WORKBOOK, OUTPUT_WORKBOOK)
    print(f"Platinum Extraction Complete. Output: {OUTPUT_WORKBOOK}")

if __name__ == "__main__":
    main()
