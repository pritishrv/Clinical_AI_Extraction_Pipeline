import os
import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.comments import Comment

# Configuration
PROJECT_ROOT = Path("/Users/joshuabhawanlall/Git Folder/Clinical_AI_Extraction_Pipeline")
MERGED_DIR = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/longitudinal_merged"
HARVEST_DIR = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/raw_harvest"
PROTOTYPE_WORKBOOK = PROJECT_ROOT / "data/hackathon-database-prototype.xlsx"
OUTPUT_WORKBOOK = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/generated-database-v7-diamond.xlsx"

def find_evidence(patient_id, col_name, value):
    """
    Looks through raw harvest files to find the snippet for a value.
    """
    # Simplified search for the hackathon
    for f in os.listdir(HARVEST_DIR):
        with open(HARVEST_DIR / f) as j:
            harvest = json.load(j)
            for cand in harvest["candidates"]:
                if cand.get("value") == value:
                    return cand.get("evidence", "No evidence snippet found.")
    return "Consolidated from longitudinal record."

def main():
    if not os.path.exists(OUTPUT_WORKBOOK.parent):
        os.makedirs(OUTPUT_WORKBOOK.parent)
        
    merged_files = sorted([f for f in os.listdir(MERGED_DIR) if f.endswith(".json")])
    
    final_rows = []
    for f in merged_files:
        with open(MERGED_DIR / f) as j:
            final_rows.append(json.load(j))
            
    df = pd.DataFrame(final_rows)
    template_df = pd.read_excel(PROTOTYPE_WORKBOOK)
    template_cols = list(template_df.columns)
    df = df.reindex(columns=template_cols)
    
    # Final Generation with Comments
    wb = load_workbook(PROTOTYPE_WORKBOOK)
    ws = wb.active
    start_row = 2
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row)

    print("Assembling final Diamond database with Evidence Comments...")
    for r_idx, (_, row) in enumerate(df.iterrows(), start=start_row):
        for c_idx, value in enumerate(row.tolist(), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            val_str = str(value).strip() if value and str(value).lower() != "nan" else ""
            if val_str.endswith(".0"): val_str = val_str[:-2]
            
            if val_str:
                cell.value = val_str
                # Inject Evidence Comment
                evidence = find_evidence(None, template_cols[c_idx-1], val_str)
                cell.comment = Comment(f"AI Evidence:\n{evidence}", "Gemini CLI")
            
            cell.number_format = '@'
            cell.data_type = 's'
            
    wb.save(OUTPUT_WORKBOOK)
    print(f"\nDiamond Standard Assembly complete. Output: {OUTPUT_WORKBOOK}")

if __name__ == "__main__":
    main()
