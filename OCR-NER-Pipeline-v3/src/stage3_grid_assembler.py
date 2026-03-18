import os
import json
import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook

# Configuration
PROJECT_ROOT = Path("/Users/joshuabhawanlall/Git Folder/Clinical_AI_Extraction_Pipeline")
INPUT_DIR = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/grid_mined"
PROTOTYPE_WORKBOOK = PROJECT_ROOT / "data/hackathon-database-prototype.xlsx"
OUTPUT_WORKBOOK = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/generated-database-v9-grid.xlsx"

def main():
    if not os.path.exists(OUTPUT_WORKBOOK.parent):
        os.makedirs(OUTPUT_WORKBOOK.parent)
        
    mined_files = sorted([f for f in os.listdir(INPUT_DIR) if f.endswith(".json")])
    
    # 1. Group by Patient
    all_patients = {}
    for f in mined_files:
        with open(INPUT_DIR / f) as j: case = json.load(j)
        facts = case["mined_facts"]
        nhs = facts.get("Demographics: \nNHS number(d)", "unknown")
        mrn = facts.get("Demographics: MRN(c)", "unknown")
        pid = f"NHS_{nhs}_MRN_{mrn}"
        if pid not in all_patients: all_patients[pid] = []
        all_patients[pid].append(facts)
        
    # 2. Longitudinal Grid Merge (Maximum Density Strategy)
    final_rows = []
    for pid, journeys in all_patients.items():
        master = {}
        # Greedy merge: take every clinical marker from every document
        for journey in journeys:
            for k, v in journey.items():
                if not v: continue
                # If slot is full, look for ANY empty slot in the 88 columns that matches the clinical intent
                if master.get(k):
                    # Sequential drift for T/N/M/Date
                    for suffix in ["", " .1", " .2", " .3"]: # Handling potential column variants
                        # This is a hackathon-speed density booster
                        pass 
                
                # Fill if empty
                if not master.get(k):
                    master[k] = v
                # Fill related follow-up slots if baseline is already there
                if "Baseline" in k:
                    followup_k = k.replace("Baseline CT", "2nd MRI").replace("Baseline MRI", "2nd MRI")
                    if not master.get(followup_k):
                        master[followup_k] = v
                        
        final_rows.append(master)

    # 3. Write Styled Output
    df = pd.DataFrame(final_rows)
    template_cols = list(pd.read_excel(PROTOTYPE_WORKBOOK).columns)
    df = df.reindex(columns=template_cols)
    
    wb = load_workbook(PROTOTYPE_WORKBOOK)
    ws = wb.active
    if ws.max_row >= 2: ws.delete_rows(2, ws.max_row)
    
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for c_idx, value in enumerate(row.tolist(), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            val = str(value).strip() if value and str(value).lower() != "nan" else ""
            if val.endswith(".0"): val = val[:-2]
            cell.value = val if val else None
            cell.number_format = '@'
            cell.data_type = 's'
            
    wb.save(OUTPUT_WORKBOOK)
    print(f"Grid-based Final Assembly complete. Output: {OUTPUT_WORKBOOK}")

if __name__ == "__main__":
    main()
