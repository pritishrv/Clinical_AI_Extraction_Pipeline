import os
import json
import pandas as pd
import re
from pathlib import Path
from openpyxl import load_workbook

# Configuration
PROJECT_ROOT = Path("/Users/joshuabhawanlall/Git Folder/Clinical_AI_Extraction_Pipeline")
PHASE_DIR = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/phase_classified"
MAPPED_DIR = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/mapped_harvest"
PROTOTYPE_WORKBOOK = PROJECT_ROOT / "data/hackathon-database-prototype.xlsx"
OUTPUT_WORKBOOK = PROJECT_ROOT / "OCR-NER-Pipeline-v3/output/generated-database-v8-obsidian.xlsx"

# --- SMART STATE ROUTING TABLE ---
ROUTING_MAP = {
    "Baseline CT: T(h)": {
        "PHASE_0_BASELINE": "Baseline CT: T(h)",
        "PHASE_1_RESTAGING": "2nd MRI: mrT",
        "PHASE_3_SURVEILLANCE": "12 week MRI: mrT"
    },
    "Baseline CT: N(h)": {
        "PHASE_0_BASELINE": "Baseline CT: N(h)",
        "PHASE_1_RESTAGING": "2nd MRI: mrN",
        "PHASE_3_SURVEILLANCE": "12 week MRI: mrN"
    },
    "Baseline MRI: mrT(h)": {
        "PHASE_0_BASELINE": "Baseline MRI: mrT(h)",
        "PHASE_1_RESTAGING": "2nd MRI: mrT",
        "PHASE_3_SURVEILLANCE": "12 week MRI: mrT"
    },
    "MDT (after 6 week: Decision ": {
        "PHASE_0_BASELINE": "MDT (after 6 week: Decision ",
        "PHASE_1_RESTAGING": "MDT (after 6 week: Decision ",
        "PHASE_3_SURVEILLANCE": "MDT (after 12 week): Decision "
    }
}

def main():
    if not os.path.exists(OUTPUT_WORKBOOK.parent): os.makedirs(OUTPUT_WORKBOOK.parent)
    
    # 1. Load MAPPED data (Density Source)
    mapped_files = sorted([f for f in os.listdir(MAPPED_DIR) if f.endswith(".json")])
    
    all_docs = []
    for f in mapped_files:
        with open(MAPPED_DIR / f) as j:
            mapped_data = json.load(j)["mapped_columns"]
            
        # Match with Phase info
        phase_file = PHASE_DIR / f.replace("_mapped.json", "_phased.json")
        phase = "PHASE_0_BASELINE"
        if os.path.exists(phase_file):
            with open(phase_file) as j:
                phase = json.load(j)["phase"]
        
        # Get patient ID
        nhs = mapped_data.get("Demographics: \nNHS number(d)", "unknown").replace(".0", "")
        all_docs.append({
            "patient_id": f"NHS_{nhs}",
            "phase": phase,
            "data": mapped_data
        })
        
    df_docs = pd.DataFrame(all_docs)
    patients = df_docs.groupby("patient_id")
    
    final_rows = []
    for pid, group in patients:
        master_record = {}
        for _, doc in group.iterrows():
            phase = doc["phase"]
            case_data = doc["data"]
            
            for col, val in case_data.items():
                if not val: continue
                
                # A. Routing
                routed = False
                if col in ROUTING_MAP:
                    target = ROUTING_MAP[col].get(phase)
                    if target:
                        master_record[target] = val
                        routed = True
                
                # B. Cumulative Greedy Fallback
                if not routed:
                    # Fill if empty or if new unique data point
                    if col not in master_record or not master_record[col]:
                        master_record[col] = val
                        
        final_rows.append(master_record)

    df_final = pd.DataFrame(final_rows)
    template_cols = list(pd.read_excel(PROTOTYPE_WORKBOOK).columns)
    df_final = df_final.reindex(columns=template_cols)
    
    # Styled Assembly
    wb = load_workbook(PROTOTYPE_WORKBOOK)
    ws = wb.active
    if ws.max_row >= 2: ws.delete_rows(2, ws.max_row)
    for r_idx, (_, row) in enumerate(df_final.iterrows(), start=2):
        for c_idx, value in enumerate(row.tolist(), start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            val = str(value).strip() if value and str(value).lower() != "nan" else ""
            if val.endswith(".0"): val = val[:-2]
            cell.value = val if val else None
            cell.number_format = '@'
            cell.data_type = 's'
    wb.save(OUTPUT_WORKBOOK)
    print(f"Obsidian Hybrid Linker complete. Output: {OUTPUT_WORKBOOK}")

if __name__ == "__main__":
    main()
