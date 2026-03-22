import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import warnings

# Configuration
INPUT_FILE = Path("v5_VLM_LLM_DirectLogic/output/hackathon_EXPERT_REFINED.xlsx")
OUTPUT_FILE = Path("v5_VLM_LLM_DirectLogic/output/hackathon_EXPERT_REFINED_CLEAN.xlsx")

def main():
    print(f"Cleaning placeholder strings from {INPUT_FILE}...")
    
    # 1. Load the workbook with openpyxl to preserve styles
    wb = load_workbook(INPUT_FILE)
    ws = wb.active
    
    # List of strings to consider as "Blank"
    placeholders = [
        'null', 'none', 'not provided', 'not recorded', 'not specified', 
        'not mentioned', 'not reported', 'unknown', 'nan', 'n/a', '-', '.', 
        'not documented', 'not available', 'no information provided'
    ]
    
    cleaned_count = 0
    
    # 2. Iterate through all cells
    for row in ws.iter_rows(min_row=2): # Skip header
        for cell in row:
            if cell.value:
                val_str = str(cell.value).strip().lower()
                
                # Check if value matches any placeholder or flexible null phrase
                if val_str in placeholders or val_str == "" or "this is null" in val_str:
                    cell.value = None
                    # IMPORTANT: Clear the fill if we clear the value
                    cell.fill = PatternFill(fill_type=None)
                    cleaned_count += 1
                # Check for "not provided" etc inside sentences? 
                # Better to be safe and only clear if it's the WHOLE value.
    
    # 3. Save the cleaned file
    wb.save(OUTPUT_FILE)
    print(f"Cleaned {cleaned_count} placeholder cells.")
    print(f"Final cleaned file: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
