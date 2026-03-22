import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path

# Configuration
PROTO_FILE = Path("data/hackathon-database-prototype.xlsx")
FINAL_FILE = Path("v5_VLM_LLM_DirectLogic/output/hackathon_EXPERT_REFINED_CLEAN.xlsx")

def main():
    print(f"Synchronizing styles from {PROTO_FILE} to {FINAL_FILE}...")
    
    # 1. Load both workbooks
    wb_proto = load_workbook(PROTO_FILE)
    ws_proto = wb_proto.active
    
    wb_final = load_workbook(FINAL_FILE)
    ws_final = wb_final.active
    
    # 2. Map Column Name -> Style in Prototype
    # We skip 'Patient Name' in the final as it's not in the prototype
    proto_styles = {}
    for cell in ws_proto[1]:
        if cell.value:
            proto_styles[cell.value] = {
                'fill': copy.copy(cell.fill) if cell.fill else None
            }
            
    # 3. Apply styles to Final File
    applied_count = 0
    for cell in ws_final[1]:
        if cell.value in proto_styles:
            style = proto_styles[cell.value]
            if style['fill']:
                cell.fill = style['fill']
                applied_count += 1
                
    # 4. Special Case: Style 'Patient Name' (Column A)
    # Let's make it match the first Demographics column style
    demo_style = next(iter(proto_styles.values()))['fill']
    if demo_style and ws_final.cell(row=1, column=1).value == "Patient Name":
        ws_final.cell(row=1, column=1).fill = demo_style
        
    wb_final.save(FINAL_FILE)
    print(f"Successfully applied {applied_count} header styles.")

if __name__ == "__main__":
    main()
